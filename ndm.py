#import python, mysql, tkinter libraryte
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from PIL import Image, ImageTk
import sv_ttk
import mysql.connector as mysql
import datetime 
from openpyxl import Workbook


#class ndm

class ndm:
    def __init__(self,root):
        self.root = root
        self.root.title("Natural Dairy Mart")
        self.root.geometry("1000x700")


        self.style = ttk.Style(self.root)
        self.style.configure('TNotebook.Tab', font=('Helvetica', '15'))
        self.style.configure('TNotebook', tabposition='n')

        current_theme =self.style.theme_use()
        self.style.theme_settings(current_theme, {"TNotebook.Tab": {"configure": {"padding": [20, 5]}}})  

        #entry styling
        self.style.configure('TEntry', font=('Helvetica', '15'))
        self.style.map('TEntry', background=[('disabled', 'black')])
        self.cart=[]

 

        sv_ttk.set_theme('light')

        #style for buttons
        self.style.configure('TButton', font=('Calibri', '15'))
        self.style.map('TButton', background=[('active', 'black')])
        self.style.map('TButton', foreground=[('active', 'blue')])

        #style for labels
        self.style.configure('TLabel', font=('Calibri', '15'))
        self.style.map('TLabel', background=[('active', 'black')])
        self.style.map('TLabel', foreground=[('active', 'white')])

        #style for entry box width
        self.style.configure('TEntry', font=('Calibri', '15'))
        self.style.map('TEntry', background=[('active', 'black')])
        self.style.map('TEntry', foreground=[('active', 'white')])
        self.style.map('TEntry', width=[('active', '100')])
        self.style.map('TEntry', relief=[('active', 'sunken')])

        #login page
        self.login_page()

    def login_page(self):
        #clear 
        for i in self.root.winfo_children():
            i.destroy()

        #ndm_bg
        self.bg=Image.open("src/img/ndm_bg.png")
        self.bg=self.bg.resize((1000,700), Image.LANCZOS)
        self.bg=ImageTk.PhotoImage(self.bg)
        self.bg_label=Label(self.root,image=self.bg)
        self.bg_label.image=self.bg
        self.bg_label.place(x=0,y=0)


        # self.login_frame = Frame(self.root)
        # self.login_frame.place(x=250, y=150, width=1000, height=700)

        self.login_label = Label(self.root, text="Login", font=("times new roman", 30,'bold'), bg="white")
        self.login_label.place(x=450, y=150)

        #label '-------------------------"
        self.line_label = Label(self.root, text="-----------------------------------------------", font=("times new roman", 20), bg="white")
        self.line_label.place(x=300, y=200)

        self.username_label = Label(self.root, text="Username", font=("times new roman", 20), bg="white")
        self.username_label.place(x=300, y=250)

        self.password_label = Label(self.root, text="Password", font=("times new roman", 20), bg="white")
        self.password_label.place(x=300, y=300)

        self.username_entry = Entry(self.root, font=("times new roman", 20), width=20)
        self.username_entry.place(x=500, y=250)

        self.password_entry = Entry(self.root,show="*", font=("times new roman", 20), width=20)
        self.password_entry.place(x=500, y=300)

        self.line_label = Label(self.root, text="-------------------------------------------------", font=("times new roman", 20), bg="white")
        self.line_label.place(x=300, y=400)

        self.login_button = Button(self.root, text="Login", font=("times new roman", 20), command=self.login,bg="#339ab0",fg="white",width=15)
        self.login_button.place(x=400, y=350)

        #label dont have an account
        self.dont_have_account_label = Label(self.root, text="Don't have an account?", font=("times new roman", 20), bg="white")
        self.dont_have_account_label.place(x=320, y=440)

        #reset my password
        self.reset_password_button = Button(self.root, text="Reset Password", font=("times new roman", 20), command=self.reset_password,fg="blue")
        self.reset_password_button.place(x=600, y=500)

        self.register_button = Button(self.root, text="Register", font=("times new roman", 20), command=self.register_page,fg="blue")
        self.register_button.place(x=600, y=440)

        #forgot password ? label
        self.forgot_password_label = Label(self.root, text="Forgot Password?", font=("times new roman", 20), bg="white")
        self.forgot_password_label.place(x=320, y=500)

    #reset password
    def reset_password(self):
        #clear
        for i in self.root.winfo_children():
            i.destroy()

        #ndm_bg
        self.bg=Image.open("src/img/ndm_bg1.png")
        self.bg=self.bg.resize((1000,700), Image.LANCZOS)
        self.bg=ImageTk.PhotoImage(self.bg)
        self.bg_label=Label(self.root,image=self.bg)
        self.bg_label.image=self.bg
        self.bg_label.place(x=0,y=0)

        self.reset_password_frame = Frame(self.root)
        self.reset_password_frame.place(x=250, y=50, width=550, height=500)

        #configure color to white
        self.reset_password_frame.configure(bg="white")
        self.reset_password_label = Label(self.reset_password_frame, text="Reset Password", font=("times new roman", 30,'bold'), bg="white")
        self.reset_password_label.place(x=100, y=00)

        #line label
        self.line_label = Label(self.reset_password_frame, text="-------------------------------------------", font=("times new roman", 20), bg="white")
        self.line_label.place(x=100, y=50)

        self.email_label = Label(self.reset_password_frame, text="Email", font=("times new roman", 20), bg="white")
        self.email_label.place(x=100, y=100)

        self.email_entry = Entry(self.reset_password_frame, font=("times new roman", 20), width=20)
        self.email_entry.place(x=100, y=150)

        #password label and entry
        self.password_label = Label(self.reset_password_frame, text="Password", font=("times new roman", 20), bg="white")
        self.password_label.place(x=100, y=200)

        self.password_entry = Entry(self.reset_password_frame, show="*",font=("times new roman", 20), width=20)
        self.password_entry.place(x=100, y=250)

        #confirm password label and entry
        self.confirm_password_label = Label(self.reset_password_frame, text="Confirm Password", font=("times new roman", 20), bg="white")
        self.confirm_password_label.place(x=100, y=300)

        self.confirm_password_entry = Entry(self.reset_password_frame, show="*", font=("times new roman", 20), width=20)
        self.confirm_password_entry.place(x=100, y=350)

        #reset password button
        self.reset_password_button = Button(self.reset_password_frame, text="Reset Password", font=("times new roman", 20), command=self.reset_password_db,bg='#339ab0')
        self.reset_password_button.place(x=100, y=400)

        #back button
        self.back_button = Button(self.reset_password_frame, text="Back", font=("times new roman", 20), command=self.login_page,bg='white',fg='blue')
        self.back_button.place(x=400, y=400)

    #reset_password_db
    def reset_password_db(self):
        email = self.email_entry.get()
        password = self.password_entry.get()
        confirm_password = self.confirm_password_entry.get()

        #validate email
        if email == "":
            messagebox.showerror("Error", "Please enter email")
        #validate password
        elif password == "":
            messagebox.showerror("Error", "Please enter password")
        #confirm password
        elif confirm_password == "":
            messagebox.showerror("Error", "Please enter confirm password")
        
        #password validation
        elif len(password) < 8:
            messagebox.showerror("Error", "Password must be at least 8 characters long")
        #email validation should have split by @ and . and should have 2 parts
        elif email.count("@")!= 1 or email.count(".")!= 1:
            messagebox.showerror("Error", "Invalid email")
        #password validation should match confirm password
        elif password != confirm_password:
            messagebox.showerror("Error", "Passwords do not match")
        else:
            mydb = mysql.connect(host="localhost", user="root", password="Kommineni@2000",port=3307,database="ndm")
            mycursor = mydb.cursor()
            #check if email exists
            mycursor.execute("SELECT * FROM user WHERE email=%s", (email,))
            row = mycursor.fetchone()
            if row:
                #update password
                mycursor.execute("UPDATE user SET password=%s WHERE email=%s", (password, email))
                mydb.commit()
                messagebox.showinfo("Success", "Password reset successful")
                self.login_page()
            else:
                messagebox.showerror("Error", "Email does not exist")



    #login
    def login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()

        #validate from database
        #if admin the admin page
        if username == "admin" and password == "admin":
            self.admin_page()
            return

        mydb = mysql.connect(host="localhost",
                                user="root",
                                password="Kommineni@2000",port=3307,
                                database="ndm")
        cursor = mydb.cursor()
        cursor.execute("select * from user where email = %s and password = %s", (username, password))
        row = cursor.fetchone()
        if row:
            self.user_id = row[0]
            self.email=row[3]

                   #collect item ids from database with user id
            self.conn = mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307,database="ndm")
            self.cursor = self.conn.cursor()
            self.sql = "SELECT * from cart where user_id=%s and status!=%s"
            self.val = (self.user_id,'ordered',)
            self.cursor.execute(self.sql, self.val)
            self.rows = self.cursor.fetchall()
            #insert into cart
            self.cart = []
            for row in self.rows:
                self.cart.append(row)



            messagebox.showinfo("Success", "Login Successful")
            self.home_page()
        else:
            messagebox.showinfo("Error", "Login Failed")
            self.username_entry.delete(0, END)

        mydb.close()


    #register
    def register_page(self):
        
        for i in self.root.winfo_children():
            i.destroy()

        #ndm_bg
        self.bg=Image.open("src/img/ndm_bg1.png")
        self.bg=self.bg.resize((1000,700), Image.LANCZOS)
        self.bg=ImageTk.PhotoImage(self.bg)
        self.bg_label=Label(self.root,image=self.bg)
        self.bg_label.image=self.bg
        self.bg_label.place(x=0,y=0)




        self.frame = Frame(self.root)   
        self.frame.place(x=250, y=50, width=550, height=500)

        #configure color to white
        self.frame.configure(bg="white")

        self.register_label = Label(self.frame, text="Register", font=("times new roman", 30,'bold'), bg="white")
        self.register_label.place(x=200, y=0)

        #label '-------------------------"
        self.line_label = Label(self.frame, text="-------------------------------------", font=("times new roman", 20), bg="white")
        self.line_label.place(x=100, y=50)

        #First Name Label
        self.first_name_label = Label(self.frame, text="First Name", font=("times new roman", 20), bg="white")
        self.first_name_label.place(x=100, y=100)

        #entry
        self.first_name_entry = Entry(self.frame)
        self.first_name_entry.place(x=350, y=100)

        #last Name Label
        self.last_name_label = Label(self.frame, text="Last Name", font=("times new roman", 20), bg="white")
        self.last_name_label.place(x=100, y=150)

        #entry
        self.last_name_entry = Entry(self.frame)
        self.last_name_entry.place(x=350, y=150)

        #email Label
        self.email_label = Label(self.frame, text="Email", font=("times new roman", 20), bg="white")
        self.email_label.place(x=100, y=200)

        #entry
        self.email_entry = Entry(self.frame)
        self.email_entry.place(x=350, y=200)

        #phonenumber Label
        self.phonenumber_label = Label(self.frame, text="Phone Number", font=("times new roman", 20), bg="white")
        self.phonenumber_label.place(x=100, y=250)

        #entry
        self.phonenumber_entry = Entry(self.frame)
        self.phonenumber_entry.place(x=350, y=250)

        #password Label
        self.password_label = Label(self.frame, text="Password", font=("times new roman", 20), bg="white")
        self.password_label.place(x=100, y=300)

        #entry
        self.password_entry = Entry(self.frame, show="*")
        self.password_entry.place(x=350, y=300)

        #confirm password Label
        self.confirm_password_label = Label(self.frame, text="Confirm Password", font=("times new roman", 20), bg="white")
        self.confirm_password_label.place(x=100, y=350)

        #entry
        self.confirm_password_entry = Entry(self.frame, show="*")
        self.confirm_password_entry.place(x=350, y=350)

        self.register_button = Button(self.frame, text="Register", font=("times new roman", 20), command=self.register,bg="#339ab0",fg="white")
        self.register_button.place(x=200, y=400)

        #already have an account
        self.already_have_account_label = Label(self.frame, text="Already have an account?", font=("times new roman", 10), bg="white")
        self.already_have_account_label.place(x=380, y=380)

        self.login_button = Button(self.frame, text="Login", font=("times new roman", 20), command=self.login_page,fg="blue")
        self.login_button.place(x=400, y=400)

    #register
    def register(self):
        first_name = self.first_name_entry.get()
        last_name = self.last_name_entry.get()
        email = self.email_entry.get()
        phonenumber = self.phonenumber_entry.get()
        password = self.password_entry.get()
        confirm_password = self.confirm_password_entry.get()

        #validate data
        if first_name == "" or last_name == "" or email == "" or phonenumber == "" or password == "" or confirm_password == "":
            messagebox.showerror("Error", "Please fill in all fields")

         #email validation should have split by @ and . and should have 2 parts
        elif email.count("@")!= 1 or email.count(".")!= 1:
            messagebox.showerror("Error", "Invalid email")
        
        #phone number validation
        elif len(phonenumber)!= 10:
            messagebox.showerror("Error", "Invalid phone number")

         #password validation
        elif len(password) < 8:
            messagebox.showerror("Error", "Password must be at least 8 characters long")

        
        elif password!= confirm_password:
            messagebox.showerror("Error", "Passwords do not match")
        
           
        else:
            mydb = mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307, database="ndm")
            mycursor = mydb.cursor()
            sql = "INSERT INTO user (firstname, lastname, email, phonenumber, password) VALUES (%s, %s, %s, %s, %s)"
            val = (first_name, last_name, email, phonenumber, password)
            mycursor.execute(sql, val)
            mydb.commit()
            messagebox.showinfo("Success", "Registration Successful")
            self.home_page()

    #home page
    def home_page(self):
        for i in self.root.winfo_children():
            i.destroy()

        #ndm_bg
        self.bg = Image.open("src/img/ndm_bg2.png")
        self.bg = self.bg.resize((1000,700), Image.LANCZOS)
        self.bg=ImageTk.PhotoImage(self.bg)
        self.bg_label=Label(self.root,image=self.bg)
        self.bg_label.image=self.bg
        self.bg_label.place(x=0,y=0)

        #set geometry
        self.root.geometry("1000x700")

        self.home_label = Label(self.root, text="Home", font=("times new roman", 30),bg="#339ab0",fg="white")
        self.home_label.place(x=450, y=10)

        #logout button left corner
        self.logout_button = Button(self.root, text="Logout", font=("times new roman", 20), command=self.logout)
        self.logout_button.place(x=50, y=10)

        #cart button left corner
        self.cart_button = Button(self.root, text="Cart", font=("times new roman", 20), command=self.cart_page)
        self.cart_button.place(x=850, y=10)

        #frame for notebook
        self.notebook_frame = ttk.Frame(self.root, width=900, height=600)
        self.notebook_frame.place(x=50, y=70)

        #notebook
        self.notebook = ttk.Notebook(self.notebook_frame, width=900, height=600)
        self.notebook.place(x=0, y=0)

        #Category tab
        self.category_tab = ttk.Frame(self.notebook, width=900, height=600)
        self.notebook.place(x=0, y=0)

        self.notebook.add(self.category_tab, text="Category")

        
        #best sellers tab
        self.best_sellers_tab = ttk.Frame(self.notebook, width=900, height=600)
        self.notebook.place(x=0, y=0)

        self.notebook.add(self.best_sellers_tab, text="Best Sellers")

        #frame inside best sellers tab
        self.best_sellers_frame = ttk.Frame(self.best_sellers_tab, width=900, height=600)
        self.best_sellers_frame.pack(expand=1, fill=BOTH)

        #get all orders data from the database
        mydb = mysql.connect(host="localhost", user="root", password="Kommineni@2000",port=3307, database="ndm")
        mycursor = mydb.cursor()
        sql = "SELECT item_id, SUM(quantity) FROM orders GROUP BY item_id ORDER BY SUM(quantity) DESC"
        mycursor.execute(sql)
        result = mycursor.fetchall()
        print(result)
        for i, item in enumerate(result[:5]):
            item_id = item[0]
            #get item name
            sql = "SELECT name FROM items WHERE id=%s"
            val = (item_id,)
            mycursor.execute(sql, val)
            item_name = mycursor.fetchone()[0]
            #get item image path
            sql = "SELECT image FROM items WHERE id=%s"
            val = (item_id,)
            mycursor.execute(sql, val)
            item_image_path = mycursor.fetchone()[0]

            image = Image.open(item_image_path)
            image = image.resize((150, 200), Image.LANCZOS)
            photo = ImageTk.PhotoImage(image)
            row = i // 5
            col = i % 5

            item_button = Button(self.best_sellers_frame, image=photo, command=lambda item=item: self.show_details(item))
            item_button.image = photo
            item_button.grid(row=row, column=col, padx=5, pady=5)
            item_label = Label(self.best_sellers_frame, text=item_name, font=("times new roman", 10), bg="#d9d9d9")
            item_label.grid(row=row+1, column=col, padx=5, pady=5)
        


        #My Orders tab
        self.my_orders_tab = ttk.Frame(self.notebook, width=900, height=600)
        self.notebook.place(x=0, y=0)
        self.notebook.add(self.my_orders_tab, text="My Orders")

        #frame inside category tab
        self.category_frame = ttk.Frame(self.category_tab, width=900, height=600)
        self.category_frame.place(x=0, y=0)

        #notebook inside category frame
        self.category_notebook = ttk.Notebook(self.category_frame, width=900, height=600)
        self.category_notebook.place(x=0, y=0)

        #get catergories from the database
        mydb = mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307,database="ndm")
        mycursor = mydb.cursor()
        sql = "SELECT distinct(category) FROM items"
        mycursor.execute(sql)
        result = mycursor.fetchall()
        for row in result:
            #tab for each category
            self.category_tab = ttk.Frame(self.category_notebook, width=900, height=600)
            #place
            self.category_notebook.place(x=0, y=0)
            self.category_notebook.add(self.category_tab, text=row[0])
            #frame inside category tab
            self.category_frame = ttk.Frame(self.category_tab, width=900, height=600)
            self.category_frame.place(x=0, y=0)

            #get it's corresponding items
            sql = "SELECT * FROM items WHERE category = %s"
            val = (row[0],)
            mycursor.execute(sql, val)
            result = mycursor.fetchall()


            #display all items
            current=0
            for i, item in enumerate(result):
                item = list(item)
                item = {
                    "item_id": item[0],
                    "item_name": item[1],
                    "item_image_path": item[5]
                }
                image = Image.open(item["item_image_path"])
                image = image.resize((150, 200), Image.LANCZOS)
                photo = ImageTk.PhotoImage(image)
                row = i // 5
                col = i % 5

                item_button = Button(self.category_frame, image=photo, command=lambda item=item: self.show_details_user(item))
                item_button.image = photo
                item_button.grid(row=row, column=col, padx=5, pady=5)
                item_label = Label(self.category_frame, text=item["item_name"], font=("times new roman", 10),bg="#d9d9d9")
                item_label.grid(row=row+1, column=col, padx=5, pady=5)

                current = i

        

        #frame for my orders
        self.my_orders_frame = ttk.Frame(self.my_orders_tab, width=900, height=600)
        self.my_orders_frame.pack(expand=1, fill=BOTH)

        #treeview for my orders
        self.my_orders_treeview = ttk.Treeview(self.my_orders_frame, columns=("item_name", "Quantity", "Price","Date Placed"), show="headings")
        self.my_orders_treeview.heading("item_name", text="Item Name")
        self.my_orders_treeview.heading("Quantity", text="Quantity")
        self.my_orders_treeview.heading("Price", text="Price")
        self.my_orders_treeview.heading("Date Placed", text="Date Placed")
        self.my_orders_treeview.pack(fill="both", expand=True)

        #column width
        self.my_orders_treeview.column("item_name", width=100)
        self.my_orders_treeview.column("Quantity", width=100)
        self.my_orders_treeview.column("Price", width=100)
        self.my_orders_treeview.column("Date Placed", width=100)
        self.my_orders_treeview.pack(fill="both", expand=True)

        #fill the data
        mydb = mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307,database="ndm")
        mycursor = mydb.cursor()
        sql = "SELECT * FROM orders WHERE user_id=%s"
        val = (self.user_id,)
        mycursor.execute(sql, val)
        result = mycursor.fetchall()
        for i in result:
            item_id = i[3]
            #get item name
            sql = "SELECT name FROM items WHERE id=%s"
            val = (item_id,)
            mycursor.execute(sql, val)
            item_name = mycursor.fetchone()[0]
            self.my_orders_treeview.insert("", "end", values=(item_name, i[4], i[5], i[6]))


    #show_details_user
    def show_details_user(self, item):
        #get item details from the database
        mydb = mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307, database="ndm")
        mycursor = mydb.cursor()
        sql = "SELECT * FROM items WHERE id = %s"
        val = (item["item_id"],)
        mycursor.execute(sql, val)
        result = mycursor.fetchall()
        print(result)
        item = [i for i in result[0]]
        self.item = {
            "item_id": item[0],
            "item_name": item[1],
            "price": item[2],
            "description": item[3],
            "category": item[4],
            "item_image_path": item[5]
        }

        print(self.item)

        #clear
        for i in self.root.winfo_children():
            i.destroy()

        #items frame
        self.item_frame = ttk.Frame(self.root, width=900, height=600)
        self.item_frame.place(x=50, y=70)

        
        #show item details label
        self.details_label = Label(self.item_frame, text="Item Details", font=("times new roman", 20),bg="#d9d9d9")
        self.details_label.place(x=50, y=50)

        #back button
        self.back_button = Button(self.item_frame, text="Back", font=("times new roman", 20), command=self.home_page, bg="#d9d9d9")
        self.back_button.place(x=800, y=50)

        #show item image
        image = Image.open(self.item["item_image_path"])
        image = image.resize((100, 150), Image.LANCZOS)
        photo = ImageTk.PhotoImage(image)
        self.item_image = Label(self.item_frame, image=photo)
        self.item_image.image = photo
        self.item_image.place(x=50, y=150)

        #show item on left
        #Name : label
        Label(self.item_frame, text="Name", font=("times new roman", 20), bg="#d9d9d9").place(x=300, y=100)
        #Price : label
        Label(self.item_frame, text="Price", font=("times new roman", 20), bg="#d9d9d9").place(x=300, y=150)
        #Description : label
        Label(self.item_frame, text="Description", font=("times new roman", 20), bg="#d9d9d9").place(x=300, y=200)
        #Category : label
        Label(self.item_frame, text="Category", font=("times new roman", 20), bg="#d9d9d9").place(x=300, y=250)
        #Quantity : combobox
        Label(self.item_frame, text="Quantity", font=("times new roman", 20), bg="#d9d9d9").place(x=300, y=300)

        self.item_name_label = Label(self.item_frame, text=self.item["item_name"], font=("times new roman", 20), bg="#d9d9d9")
        self.item_name_label.place(x=500, y=100)
        self.item_price_label = Label(self.item_frame, text=self.item["price"], font=("times new roman", 20), bg="#d9d9d9")
        self.item_price_label.place(x=500, y=150)
        self.item_description_label = Label(self.item_frame, text=self.item["description"], font=("times new roman", 20), bg="#d9d9d9")
        self.item_description_label.place(x=500, y=200)
        self.item_category_label = Label(self.item_frame, text=self.item["category"], font=("times new roman", 20), bg="#d9d9d9")
        self.item_category_label.place(x=500, y=250)

        #Quantity combo box
        list=[1,2,3,4,5]
        self.quantity_combobox = ttk.Combobox(self.item_frame, values=list, state="readonly")
        self.quantity_combobox.current(0)
        self.quantity_combobox.place(x=500, y=300)

        #bind event
        self.quantity_combobox.bind("<<ComboboxSelected>>", self.quantity_combobox_selected)

        #total price label
        self.total_price_label = Label(self.item_frame, text="Total Price", font=("times new roman", 20), bg="#d9d9d9")
        self.total_price_label.place(x=500, y=350)
        self.total_price = Label(self.item_frame, text=self.item["price"], font=("times new roman", 20), bg="#d9d9d9")
        self.total_price.place(x=500, y=400)

        #add to cart button
        self.add_to_cart_button = Button(self.item_frame, text="Add to Cart", font=("times new roman", 20), command=lambda item=item: self.add_to_cart(item), bg="#d9d9d9")
        self.add_to_cart_button.place(x=300, y=500)

       

    #quantity_combobox_selected
    def quantity_combobox_selected(self, event):
        self.total_price.config(text=round(float(self.item["price"]),2) * int(self.quantity_combobox.get()))

        #place order button
        self.place_order_button.config(state="normal")

    #add_to_cart
    def add_to_cart(self, itm):

        print(itm)

        # add item to cart
        #insert into cart table
        #create connection
        self.connection=mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307, database="ndm")
        self.cursor=self.connection.cursor()
        self.cursor.execute("INSERT INTO cart (user_id,item_id, quantity) VALUES (%s, %s,%s)", (self.user_id,itm[0], int(self.quantity_combobox.get())))
        self.connection.commit()

        #get all items in cart
        self.cursor.execute("SELECT * FROM cart WHERE user_id=%s and status!=%s", (self.user_id,'ordered',))
        self.row = self.cursor.fetchall()

        #insert into cart list
        self.cart=[]
        for i in self.row:
            self.cart.append(i)

        print(self.cart)


        # added to cart message
        messagebox.showinfo("Success", "Item added to cart")
        #back to home page
        self.home_page()



    #admin_page
    def admin_page(self):
        for i in self.root.winfo_children():
            i.destroy()

        
        #ndm_bg
        self.bg = Image.open("src/img/ndm_bg2.png")
        self.bg = self.bg.resize((1000,700), Image.LANCZOS)
        self.bg=ImageTk.PhotoImage(self.bg)
        self.bg_label=Label(self.root,image=self.bg)
        self.bg_label.image=self.bg
        self.bg_label.place(x=0,y=0)

        #set geometry
        self.root.geometry("1000x700")

        self.admin_label = Label(self.root, text="Admin", font=("times new roman", 30), bg="#339ab0", fg="white")
        self.admin_label.place(x=450, y=10)


        #logout button left corner
        self.logout_button = Button(self.root, text="Logout", font=("times new roman", 20), command=self.logout)
        self.logout_button.place(x=50, y=10)


        #frame for notebook
        self.notebook_frame = ttk.Frame(self.root, width=900, height=600)
        self.notebook_frame.place(x=50, y=70)


        #frame for notebook
        self.notebook = ttk.Notebook(self.notebook_frame, width=900, height=600)
        self.notebook.place(x=0, y=0)

        #Items tab
        self.items_tab = ttk.Frame(self.notebook, width=900, height=600)
        self.notebook.place(x=0, y=0)

        self.notebook.add(self.items_tab, text="Items")

        #frame inside items tab
        self.items_frame = ttk.Frame(self.items_tab, width=900, height=600)
        self.items_frame.pack(expand=1, fill=BOTH)

        #notebook inside items frame
        self.items_notebook = ttk.Notebook(self.items_frame, width=900, height=600)
        self.items_notebook.pack(expand=1, fill=BOTH)


        #collect all the categories from the table to list
        mydb = mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307, database="ndm")
        mycursor = mydb.cursor()
        sql = "SELECT distinct(category) from items"
        mycursor.execute(sql)
        rows = mycursor.fetchall()
        self.category_list = []
        for row in rows:
            self.category_list.append(row[0])
            #create frame for each category
            self.category_frame = ttk.Frame(self.items_notebook, width=900, height=600)
            #pack
            self.category_frame.pack(expand=1, fill='both')
            self.items_notebook.add(self.category_frame, text=row[0])

            #canvas inside frame
            self.category_canvas = Canvas(self.category_frame, width=900, height=600)
            self.category_canvas.place(x=0, y=0)
            self.category_canvas.create_rectangle(0, 0, 900, 600, fill="white")


            #get all items from the database
            con=mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307, database="ndm")
            mycursor = con.cursor()
            sql = "SELECT * from items where category=%s"
            val = (row[0],)
            mycursor.execute(sql, val)
            rows = mycursor.fetchall()

            current=0
            for i, item in enumerate(rows):
                item = list(item)
                item = {
                    "item_id": item[0],
                    "item_name": item[1],
                    "item_image_path": item[5]
                }
                image = Image.open(item["item_image_path"])
                image = image.resize((150, 200), Image.LANCZOS)
                photo = ImageTk.PhotoImage(image)
                row = i // 5
                col = i % 5

                item_button = Button(self.category_frame, image=photo, command=lambda item=item: self.show_details(item))
                item_button.image = photo
                item_button.grid(row=row, column=col, padx=5, pady=5)
                current = i

            
        mydb.close()

        #add item tab in notebook
        self.add_item_tab = ttk.Frame(self.notebook,width=900, height=600)
        self.notebook.place(x=0, y=0)
        self.notebook.add(self.add_item_tab, text="Add Item")

        


        #take item details
        self.item_name_label = Label(self.add_item_tab, text="Item Name", font=("times new roman", 20), bg="#d9d9d9")
        self.item_name_label.place(x=100, y=100)
        self.item_name_entry = Entry(self.add_item_tab, bg="#d9d9d9",width=28)
        self.item_name_entry.place(x=300, y=100)

        #price
        self.price_label = Label(self.add_item_tab, text="Price", font=("times new roman", 20), bg="#d9d9d9")
        self.price_label.place(x=100, y=150)
        self.price_entry = Entry(self.add_item_tab, bg="#d9d9d9",width=28)
        self.price_entry.place(x=300, y=150)

        #category
        self.category_label = Label(self.add_item_tab, text="Category", font=("times new roman", 20), bg="#d9d9d9")
        self.category_label.place(x=100, y=200)

        #categories list 
        #list
        self.category_list=['Milk', 'Eggs', 'Cheese','Butter','Yogurt', 'Cream','Icecream']
        self.category_entry = ttk.Combobox(self.add_item_tab, values=self.category_list)
        self.category_entry.place(x=300, y=200)

        #description
        self.description_label = Label(self.add_item_tab, text="Description", font=("times new roman", 20), bg="#d9d9d9")
        self.description_label.place(x=100, y=250)
        self.description_entry = Entry(self.add_item_tab, bg="#d9d9d9",width=28)
        self.description_entry.place(x=300, y=250)

        #add button that opens files and take image path
        self.take_image_button = Button(self.add_item_tab, text="Take Image", font=("times new roman", 20), command=self.take_image)
        self.take_image_button.place(x=100, y=300)

        #add button that adds item to the database
        self.add_item_button = Button(self.add_item_tab, text="Add Item", font=("times new roman", 20), command=self.add_item)
        self.add_item_button.place(x=300, y=300)

        #download reports tab
        self.download_reports_tab = ttk.Frame(self.notebook, width=900, height=600)
        self.notebook.place(x=0, y=0)
        self.notebook.add(self.download_reports_tab, text="Download Reports")

        #frame inside download reports tab
        self.download_reports_frame = ttk.Frame(self.download_reports_tab, width=900, height=600) 
        self.download_reports_frame.pack(expand=1, fill=BOTH)

        #download sales report button
        self.download_sales_report_button = Button(self.download_reports_frame, text="Download Sales Report", font=("times new roman", 20), command=self.download_sales_report_txt)
        self.download_sales_report_button.place(x=300, y=100)

        #download items report button
        self.download_items_report_button = Button(self.download_reports_frame, text="Download Items Report", font=("times new roman", 20), command=self.download_items_report)
        self.download_items_report_button.place(x=300, y=200)

        #download users report button
        self.download_users_report_button = Button(self.download_reports_frame, text="Download Users Report", font=("times new roman", 20), command=self.download_users_report)
        self.download_users_report_button.place(x=300, y=300)

    #download_sales_report
    def download_sales_report_txt(self):
        mydb = mysql.connect(host="localhost", user="root", password="Kommineni@2000",port=3307, database="ndm")
        mycursor = mydb.cursor()

        #get all the orders
        mycursor.execute("SELECT * FROM orders")
        rows = mycursor.fetchall()
        #print(rows)

        #write to a excel
        #todays'date
        today = datetime.datetime.now()
        today = today.strftime("%d-%m-%Y")
        #create a workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        #append date
        ws.append(["Date", today])
        ws.append(["S.No","Order ID", "User ID", "Item ID", "Quantity", "Price", "Date Placed"])
        for row in rows:
            ws.append([row[0], row[1], row[2], row[3], row[4], row[5],row[6]])
        wb.save("sales_report.xlsx")
        messagebox.showinfo("Success", "Sales Report Downloaded")



        # #write to a file
        # with open("sales_report.txt", "w") as file:
        #     file.write("Sales Report\n")
        #     file.write("Order ID\tUser ID\tItem ID\tQuantity\tPrice\tDate Placed\n")
        #     for row in rows:
        #         file.write(f"{row[0]}\t{row[1]}\t{row[2]}\t{row[3]}\t{row[4]}\t{row[5]}\n")
        #         file.write("\n")
        mydb.close()
        

    #download_items_report_pdf
    def download_items_report_pdf(self):
        mydb = mysql.connect(host="localhost", user="root", password="Kommineni@2000",port=3307, database="ndm")
        mycursor = mydb.cursor()
        #get all the items
        mycursor.execute("SELECT * FROM items")
        rows = mycursor.fetchall()
        #print(rows)
        #write to a file using fpdf
    
    #download_items_report
    def download_items_report(self):
        mydb = mysql.connect(host="localhost", user="root", password="Kommineni@2000",port=3307, database="ndm")
        mycursor = mydb.cursor()
        #get all the items
        mycursor.execute("SELECT * FROM items")
        rows = mycursor.fetchall()
        #print(rows)
        #write to a file excel
        #todays'date
        today = datetime.datetime.now()
        today = today.strftime("%d-%m-%Y")
        #create a workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Items Report"
        #append date
        ws.append(["Date", today])
        ws.append(["Item ID", "Item Name", "Price", "Description", "Category"])
        for row in rows:
            ws.append([row[0], row[1], row[2], row[3], row[4]])
        wb.save("items_report.xlsx")
        messagebox.showinfo("Success", "Items Report Downloaded")

        # with open("items_report.txt", "w") as file:
        #     file.write("Items Report\n")
        #     file.write("Item ID\tItem Name\tPrice\tCategory\tDescription\n")
        #     for row in rows:
        #         file.write(f"{row[0]}\t{row[1]}\t{row[2]}\t{row[3]}\t{row[4]}\n")
        #         file.write("\n")
        mydb.close()
        
    #download_users_report
    def download_users_report(self):
        mydb = mysql.connect(host="localhost", user="root", password="Kommineni@2000",port=3307, database="ndm")
        mycursor = mydb.cursor()
        #get all the users
        mycursor.execute("SELECT * FROM user")
        rows = mycursor.fetchall()
        #print(rows)
        #write to a file

        #todays'date
        today = datetime.datetime.now()
        today = today.strftime("%d-%m-%Y")
        #create a workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Users Report"
        #append date
        ws.append(["Date", today])
        ws.append(["User ID", "FirstName", "LastName", "Email", "Phone"])
        for row in rows:
            ws.append([row[0], row[1], row[2], row[3], row[4]])
        wb.save("users_report.xlsx")
        messagebox.showinfo("Success", "Users Report Downloaded")

    #show details
    def show_details(self, item):
        #item details
        self.item_id = item["item_id"]
        self.item_name = item["item_name"]

        #clear screen
        for i in self.root.winfo_children():
            i.destroy()

        #bg image
        self.bg=Image.open("src/img/ndm_bg2.png")
        self.bg=self.bg.resize((1000,700), Image.LANCZOS)
        self.bg=ImageTk.PhotoImage(self.bg)
        self.bg_label=Label(self.root, image=self.bg)
        self.bg_label.image=self.bg
        self.bg_label.place(x=0, y=0)

        #get details 
        mydb = mysql.connect(host="localhost", user="root", password="Kommineni@2000",port=3307, database="ndm")
        mycursor = mydb.cursor()
        sql = "SELECT * from items where id=%s"
        val = (self.item_id,)
        mycursor.execute(sql, val)
        rows = mycursor.fetchall()
        item = list(rows[0])

        print(item)
        self.item = {
            "item_id": item[0],
            "item_name": item[1],
            "price": item[2],
            "category": item[3],
            "description": item[4],
            "image_path": item[5]
        }

        self.root.configure(bg="#d9d9d9")

        print(self.item)
        #items detials frame
        self.item_details_frame = Frame(self.root, width=600, height=500)
        self.item_details_frame.place(x=200, y=100)

        #configure bg
        self.item_details_frame.configure(bg="white")


        #items details page label
        self.item_details_label = Label(self.item_details_frame, text="Item Details", font=("times new roman", 30,'bold'), bg="white").place(x=100, y=50)

        #display these data as labels and entries and entries filled to update
        self.item_name_label = Label(self.item_details_frame, text="Item Name", font=("times new roman", 20), bg="white").place(x=100, y=100)
        self.item_name_entry = ttk.Entry(self.item_details_frame)
        self.item_name_entry.place(x=300, y=100)

        self.price_label = Label(self.item_details_frame, text="Price", font=("times new roman", 20), bg="white")
        self.price_label.place(x=100, y=150)
        self.price_entry = ttk.Entry(self.item_details_frame)
        self.price_entry.place(x=300, y=150)
        self.category_label = Label(self.item_details_frame, text="Category", font=("times new roman", 20), bg="white")
        self.category_label.place(x=100, y=200)
        self.category_entry = ttk.Entry(self.item_details_frame)
        self.category_entry.place(x=300, y=200)
        self.description_label = Label(self.item_details_frame, text="Description", font=("times new roman", 20), bg="white")
        self.description_label.place(x=100, y=250)
        self.description_entry = ttk.Entry(self.item_details_frame,)
        self.description_entry.place(x=300, y=250)

        #fill the entry with the data
        self.item_name_entry.insert(0, self.item["item_name"])
        self.price_entry.insert(0, self.item["price"])
        self.category_entry.insert(0, self.item["description"])
        self.description_entry.insert(0, self.item["category"])

        #upload image button
        self.take_image_button = Button(self.item_details_frame, text="Take Image", font=("times new roman", 20), command=self.take_image).place(x=100, y=300)

        #update button
        self.update_button = Button(self.item_details_frame, text="Update", font=("times new roman", 20),bg='#339ab0',fg='white', command=self.update_item).place(x=300, y=300)

        #delete item
        self.delete_button = Button(self.item_details_frame, text="Delete", font=("times new roman", 20),bg='#339ab0',fg='white', command=self.delete_item).place(x=500, y=300)

        #down back to admin screen button
        self.back_button = Button(self.item_details_frame, text="Back", font=("times new roman", 20),bg='#339ab0',fg='white', command=self.admin_page).place(x=100, y=400)


    
    #delete
    def delete_item(self):
        mydb = mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307, database="ndm")
        mycursor = mydb.cursor()
        sql = "DELETE FROM items WHERE id=%s"
        val = (self.item_id,)
        mycursor.execute(sql, val)
        mydb.commit()
        messagebox.showinfo("Success", "Item Deleted")
        #back to admin page
        self.admin_page()



    #update
    def update_item(self):
        #take item details
        item_name = self.item_name_entry.get()
        price = self.price_entry.get()
        category = self.category_entry.get()
        description = self.description_entry.get()
        image_path = self.item['image_path']

        #validate data
        if item_name == "" or price == "" or category == "" or description == "" or image_path == "":
            messagebox.showerror("Error", "Please fill all the fields")
        else:
            mydb = mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307, database="ndm")
            mycursor = mydb.cursor()
            sql = "UPDATE items SET name=%s, price=%s, category=%s, description=%s, image=%s WHERE id=%s"
            val = (item_name, price, category, description, image_path, self.item_id)
            mycursor.execute(sql, val)
            mydb.commit()
            messagebox.showinfo("Success", "Item Updated")
            #back to admin page
            self.admin_page()


        #add item
    def add_item(self):
        #take item details
        item_name = self.item_name_entry.get()
        price = self.price_entry.get()
        category = self.category_entry.get()
        description = self.description_entry.get()
        image_path = self.image_path

        #validate data
        if item_name == "" or price == "" or category == "" or description == "" or image_path == "":
            messagebox.showerror("Error", "Please fill all the fields")
        else:
            mydb = mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307,database="ndm")
            mycursor = mydb.cursor()
            sql = "INSERT INTO items (name, price, category, description, image) VALUES (%s, %s, %s, %s, %s)"
            val = (item_name, price, category, description, image_path)
            mycursor.execute(sql, val)
            mydb.commit()
            messagebox.showinfo("Success", "Item Added")
            self.items_notebook.add(self.items_tab, text=category)
            #back to admin page
            self.admin_page()


    #take image
    def take_image(self):
        self.image_path = filedialog.askopenfilename()
        print(self.image_path)
        #if image path is not empty
        if self.image_path!= "":
            messagebox.showinfo("Success", "Image Uploaded")
        else:
            messagebox.showerror("Error", "No image selected")
        

    #logout
    def logout(self):
        self.login_page()
    
    #cart
    def cart_page(self):
        #clear 
        for i in self.root.winfo_children():
            i.destroy()

        #cart frame
        self.cart_frame = ttk.Frame(self.root, width=1000, height=700)
        self.cart_frame.place(x=0, y=0)

    

        


        #cart page label
        self.cart_label = Label(self.cart_frame, text="Cart Items", font=("times new roman", 20), bg="#d9d9d9").place(x=100, y=50)


        #treeview frame
        self.cart_treeview_frame = ttk.Frame(self.cart_frame, width=900, height=400)
        self.cart_treeview_frame.place(x=50, y=100,width=900, height=400)

        
        #treeview of all items in cart
        self.cart_treeview = ttk.Treeview(self.cart_treeview_frame, columns=("item_name", "Quantity", "Price"), show="headings")
        self.cart_treeview.heading("item_name", text="Item Name")
        self.cart_treeview.heading("Quantity", text="Quantity")
        self.cart_treeview.heading("Price", text="Price")
        self.cart_treeview.place(x=50, y=100,width=900, height=400)



        #column width
        self.cart_treeview.column("item_name", width=200)
        self.cart_treeview.column("Quantity", width=100)
        self.cart_treeview.column("Price", width=100)

        #scroll bar
        scrollbar = ttk.Scrollbar(self.cart_treeview_frame, command=self.cart_treeview.yview)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.cart_treeview.configure(yscrollcommand=scrollbar.set)
        self.cart_treeview.pack(fill=BOTH, expand=1)

        total = 0
        #get data from items and fill the table
        for i in self.cart:
            print(i)
            item_id = i[2]
            quantity = i[3]

            mydb = mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307,database="ndm")
            mycursor = mydb.cursor()
            sql = "SELECT * FROM items WHERE id=%s"
            val = (item_id,)
            mycursor.execute(sql, val)
            result = mycursor.fetchone()

            item_price = result[2]
            totalprice = round(float(item_price),2) * int(quantity)
            total = total + totalprice
            self.cart_treeview.insert("", "end", text=result[1], values=(result[1], quantity, totalprice))
            mycursor.close()
            mydb.close()

        #total price label
        self.total_price_label = Label(self.cart_frame, text="Total Price: " + str(total), font=("times new roman", 20), bg="#d9d9d9").place(x=100, y=500)

        

        #back button
        self.back_button = Button(self.cart_frame, text="Back", font=("times new roman", 20), command=self.home_page).place(x=900, y=30)

        #delete item cart button
        self.delete_button = Button(self.cart_frame, text="Delete", font=("times new roman", 20), command=self.delete_item_cart).place(x=50, y=600)

        #place order button
        self.place_order_button = Button(self.cart_frame, text="Place Order", font=("times new roman", 20), command=self.place_order).place(x=800, y=600)


    #place order
    def place_order(self):
        #take user details
        mydb = mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307, database="ndm")

        try:
            #get current max order id
            mycursor = mydb.cursor()
            sql = "SELECT max(order_id) FROM orders"
            mycursor.execute(sql)
            result = mycursor.fetchone()
            max_order_id = result[0]
            max_order_id = int(max_order_id) + 1
        except:
            max_order_id = 1


        #insert items into orders table
        for i in self.cart:
            item_id = i[2]
            quantity = i[3]

            
            #CONNECTION
            mydb = mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307, database="ndm")
            mycursor = mydb.cursor()

            #get item price
            mycursor = mydb.cursor()
            sql = "SELECT * FROM items WHERE id=%s"
            val = (item_id,)
            mycursor.execute(sql, val)
            result = mycursor.fetchone()
            item_price = result[2]
            totalprice = round(float(item_price),2) * int(quantity)

            #CONNECTION
            mydb = mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307, database="ndm")
            mycursor = mydb.cursor()

            sql = "INSERT INTO orders (order_id,user_id,item_id,quantity,price) VALUES (%s, %s, %s, %s, %s)"
            val = (max_order_id, self.user_id, item_id, quantity, totalprice)
            mycursor.execute(sql, val)
            mydb.commit()

            #update cart table status to ordered with item id and user id
            sql = "UPDATE cart SET status='ordered' WHERE item_id=%s AND user_id=%s"
            val = (item_id, self.user_id)
            mycursor = mydb.cursor()
            mycursor.execute(sql, val)
            mydb.commit()
            mycursor.close()

        #messagebox
        messagebox.showinfo("Success", "Order Placed")
        self.cart = []
        #back to home page

        self.home_page()
    #delete_item cart button
    def delete_item_cart(self):
        #delete item from cart list
        #take item id 
        item_name = self.cart_treeview.focus()
        item_name = self.cart_treeview.item(item_name)["values"][0]

        print(item_name)

        #get itemid from item name
        mydb = mysql.connect(host="localhost", user="root",  password="Kommineni@2000",port=3307,database="ndm")
        mycursor = mydb.cursor()
        sql = "SELECT * FROM items WHERE name=%s"
        val = (item_name,)
        mycursor.execute(sql, val)
        result = mycursor.fetchone()
        item_id = result[0]

        #delete item from cart table
        sql = "DELETE FROM cart WHERE item_id=%s AND user_id=%s"
        val = (item_id, self.user_id)
        mycursor = mydb.cursor()
        mycursor.execute(sql, val)
        mydb.commit()
        mycursor.close()
        mydb.close()
        #delete item from cart list
        for i in self.cart:
            if i[2] == item_id:
                self.cart.remove(i)
                break
        
        #reload cart
        self.cart_page()
    
#start program

app = ndm(Tk())
app.root.mainloop()